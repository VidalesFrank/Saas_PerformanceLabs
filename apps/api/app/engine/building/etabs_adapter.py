"""
ETABS Table Adapter: normaliza XLSX exportado de ETABS 23 al formato ETABS 17
que espera el motor (ImportCSIData / archetype_1.py).

Detección de versión:
  - ETABS 17 → tiene hoja "Objects and Elements - Shells"
  - ETABS 23 → tiene hoja "Objects and Elements - Areas"

Si el archivo ya es E17 (o e2k-generated), no se hace nada.
"""
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ── Helpers ───────────────────────────────────────────────────────────────────

def _read(xl: pd.ExcelFile, sheet: str) -> pd.DataFrame | None:
    """Lee una hoja con el convenio de 3 filas de encabezado y devuelve solo datos."""
    if sheet not in xl.sheet_names:
        return None
    df = pd.read_excel(xl, sheet_name=sheet, skiprows=1, header=0)
    return df.drop(index=0).reset_index(drop=True)


def _write_sheet(wb: Workbook, sheet_name: str, table_key: str,
                 columns: list, units: list, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=sheet_name)
    ws.append([table_key])
    ws.append(columns)
    ws.append(units)
    if df is not None and len(df):
        for row in dataframe_to_rows(df[columns], index=False, header=False):
            ws.append(row)


def detect_version(xlsx_path: str) -> str:
    """Devuelve '23' si el archivo es ETABS 23, '17' si es ETABS 17 (o compatible)."""
    try:
        xl = pd.ExcelFile(xlsx_path)
        if 'Objects and Elements - Areas' in xl.sheet_names:
            return '23'
    except Exception:
        pass
    return '17'


# ── Transformadores por tabla ─────────────────────────────────────────────────

def _t_joints(df: pd.DataFrame) -> pd.DataFrame:
    """Objects and Elements - Joints (E23→E17): Element Name → Element Label."""
    df = df.copy()
    df.rename(columns={'Element Name': 'Element Label'}, inplace=True)
    cols = ['Story', 'Element Label', 'Object Type', 'Object Label',
            'Global X', 'Global Y', 'Global Z']
    return df[[c for c in cols if c in df.columns]]


def _t_frames(df: pd.DataFrame) -> pd.DataFrame:
    """Objects and Elements - Frames (E23→E17): columnas de joints y label."""
    df = df.copy()
    df.rename(columns={
        'Element Name': 'Element Label',
        'Elm JtI': 'Joint I',
        'Elm JtJ': 'Joint J',
    }, inplace=True)
    cols = ['Story', 'Element Label', 'Object Type', 'Object Label',
            'Joint I', 'Joint J']
    return df[[c for c in cols if c in df.columns]]


def _t_shells(df: pd.DataFrame) -> pd.DataFrame:
    """Objects and Elements - Areas → Objects and Elements - Shells."""
    df = df.copy()
    df.rename(columns={
        'Element Name': 'Element Label',
        'Object Type':  'Area Type',
        'Object Label': 'Area Label',
        'Elm Jt1': 'Joint 1',
        'Elm Jt2': 'Joint 2',
        'Elm Jt3': 'Joint 3',
        'Elm Jt4': 'Joint 4',
    }, inplace=True)
    cols = ['Story', 'Element Label', 'Area Type', 'Area Label',
            'Joint 1', 'Joint 2', 'Joint 3', 'Joint 4']
    return df[[c for c in cols if c in df.columns]]


def _t_floor_conn(df: pd.DataFrame) -> pd.DataFrame:
    """Floor Object Connectivity → Floor Connectivity.

    E23 tiene 4 columnas de joints (UniquePt1..4); E17 tiene una columna 'Points'
    con los labels separados por ';'.
    """
    df = df.copy()
    pt_cols = [c for c in ['UniquePt1', 'UniquePt2', 'UniquePt3', 'UniquePt4'] if c in df.columns]
    def join_pts(row):
        pts = [int(row[c]) for c in pt_cols if pd.notna(row.get(c))]
        return ';'.join(str(p) for p in pts)
    df['Points'] = df.apply(join_pts, axis=1)
    cols = ['Story', 'Unique Name', 'Points', 'Perimeter', 'Area']
    return df[[c for c in cols if c in df.columns]]


def _t_materials(df_conc: pd.DataFrame,
                 df_mech: pd.DataFrame | None) -> pd.DataFrame:
    """Mat Prop - Concrete Data + Mat Prop - Basic Mech Props → Material Properties - Concrete."""
    df = df_conc.copy()
    df.rename(columns={
        'Material':   'Name',
        'LtWtConc':   'Lightweight?',
    }, inplace=True)

    # Fc: kN/m² → MPa (÷ 1000)
    df['Fc'] = pd.to_numeric(df['Fc'], errors='coerce') / 1000.0

    # Datos mecánicos
    if df_mech is not None:
        mech = df_mech.copy()
        mech.rename(columns={'Material': 'Name'}, inplace=True)
        mech['E']       = pd.to_numeric(mech.get('E1', 0), errors='coerce') / 1000.0
        mech['G']       = pd.to_numeric(mech.get('G12', 0), errors='coerce') / 1000.0
        mech['Poisson'] = pd.to_numeric(mech.get('U12', 0.2), errors='coerce')
        df = df.merge(mech[['Name', 'E', 'G', 'Poisson']], on='Name', how='left')
    else:
        df['E']       = 4700.0 * np.sqrt(df['Fc'].clip(lower=0))
        df['Poisson'] = 0.2
        df['G']       = df['E'] / (2 * (1 + df['Poisson']))

    # Normalizar Lightweight?
    df['Lightweight?'] = df['Lightweight?'].map(
        lambda v: 'No' if str(v).strip().lower() in ('no', 'nan', '') else 'Yes'
    )

    cols = ['Name', 'Fc', 'E', 'G', 'Poisson', 'Lightweight?']
    return df[[c for c in cols if c in df.columns]]


def _t_restraints(df: pd.DataFrame) -> pd.DataFrame:
    """Joint Assigns - Restraints → Joint Assignments - Restraints."""
    df = df.copy()
    df.rename(columns={'UniqueName': 'Unique Name'}, inplace=True)
    cols = ['Story', 'Unique Name', 'UX', 'UY', 'UZ', 'RX', 'RY', 'RZ']
    return df[[c for c in cols if c in df.columns]]


def _t_frame_sections(df: pd.DataFrame) -> pd.DataFrame:
    """Frame Prop - Summary → Frame Sections.

    E23 no tiene t3/t2. Para secciones rectangulares:
      t3 = sqrt(12 * I33 / Area)
      t2 = sqrt(12 * I22 / Area)
    """
    df = df.copy()
    for col in ['Area', 'I33', 'I22', 'J']:
        df[col] = pd.to_numeric(df.get(col, 0), errors='coerce').fillna(0)

    def calc_t(i_val, area):
        if area > 1e-12:
            v = 12 * i_val / area
            return math.sqrt(max(v, 0))
        return 0.0

    df['t3'] = df.apply(lambda r: calc_t(r['I33'], r['Area']), axis=1)
    df['t2'] = df.apply(lambda r: calc_t(r['I22'], r['Area']), axis=1)

    # E17 exporta Area en cm² e I/J en cm⁴ — convertir desde m²/m⁴
    df['Area'] = df['Area'] * 1e4   # m² → cm²
    df['I33']  = df['I33']  * 1e8   # m⁴ → cm⁴
    df['I22']  = df['I22']  * 1e8
    if 'J' in df.columns:
        df['J'] = df['J'] * 1e8

    cols = ['Name', 'Material', 'Shape', 't3', 't2', 'Area', 'I33', 'I22', 'J']
    return df[[c for c in cols if c in df.columns]]


def _t_frame_assigns(df_assigns: pd.DataFrame,
                     df_summary: pd.DataFrame | None) -> pd.DataFrame:
    """Frame Assigns - Sect Prop → Frame Assignments - Sections.

    Design Type viene directamente de Frame Assigns - Summary (Beam/Column).
    """
    df = df_assigns.copy()
    df.rename(columns={
        'UniqueName':        'Unique Name',
        'Section Property':  'Analysis Section',
        'Shape':             'Section Type',
    }, inplace=True)
    df['Design Section'] = df['Analysis Section']

    if df_summary is not None:
        type_map = (
            df_summary[['Story', 'Label', 'Design Type']]
            .drop_duplicates(subset=['Story', 'Label'])
        )
        df = df.merge(type_map, on=['Story', 'Label'], how='left')
        df['Design Type'] = df['Design Type'].fillna('Beam')
    else:
        df['Design Type'] = 'Beam'

    cols = ['Story', 'Label', 'Unique Name', 'Analysis Section',
            'Design Section', 'Design Type', 'Section Type']
    return df[[c for c in cols if c in df.columns]]


def _t_shell_assigns(df: pd.DataFrame) -> pd.DataFrame:
    """Area Assigns - Sect Prop → Shell Assignments - Sections."""
    df = df.copy()
    df.rename(columns={
        'UniqueName':       'Unique Name',
        'Section Property': 'Section',
    }, inplace=True)
    cols = ['Story', 'Unique Name', 'Section']
    return df[[c for c in cols if c in df.columns]]


def _t_slab_sections(df: pd.DataFrame) -> pd.DataFrame:
    """Slab Property Definitions → Shell Sections - Slab."""
    df = df.copy()
    # Normalizar nombre de columna One-Way
    ow_src = next((c for c in df.columns if 'one' in c.lower() and 'way' in c.lower()), None)
    if ow_src:
        df.rename(columns={ow_src: 'One-Way Load Distribution?'}, inplace=True)
    cols = ['Name', 'Material', 'Slab Thickness', 'One-Way Load Distribution?']
    return df[[c for c in cols if c in df.columns]]


def _t_wall_sections(df: pd.DataFrame) -> pd.DataFrame:
    """Wall Property Def - Specified → Shell Sections - Wall."""
    df = df.copy()
    df.rename(columns={'Wall Thickness': 'Thickness'}, inplace=True)
    cols = ['Name', 'Material', 'Thickness']
    return df[[c for c in cols if c in df.columns]]


def _t_shell_loads(df: pd.DataFrame) -> pd.DataFrame:
    """Area Loads - Uniform → Shell Loads - Uniform."""
    df = df.copy()
    df.rename(columns={'UniqueName': 'Unique Name'}, inplace=True)
    cols = ['Story', 'Label', 'Unique Name', 'Load Pattern', 'Direction', 'Load']
    return df[[c for c in cols if c in df.columns]]


def _t_frame_loads(df: pd.DataFrame) -> pd.DataFrame:
    """Frame Loads - Distributed: normaliza columnas E23 → E17."""
    df = df.copy()
    df.rename(columns={
        'UniqueName':            'Unique Name',
        'Relative Distance A':   'Distance A',
        'Relative Distance B':   'Distance B',
        'Force A':               'Load A',
        'Force B':               'Load B',
    }, inplace=True)
    # Alias para compatibilidad con el engine (usa 'Force at End')
    if 'Load A' in df.columns and 'Force at End' not in df.columns:
        df['Force at End'] = df['Load A']
    # Normalizar Direction: 'Global-Z' → 'Gravity' para cargas de gravedad
    if 'Direction' in df.columns:
        df['Direction'] = df['Direction'].map(
            lambda v: 'Gravity' if str(v).strip().lower() in ('global-z', '-z', 'gravity', 'z') else v
        )
    cols = ['Story', 'Label', 'Unique Name', 'Load Pattern', 'Direction',
            'Distance A', 'Distance B', 'Load A', 'Load B', 'Force at End']
    return df[[c for c in cols if c in df.columns]]


def _t_col_rebar(df: pd.DataFrame) -> pd.DataFrame:
    """Frame Sec Def - Conc Col Reinf → Concrete Column Rebar Data."""
    df = df.copy()
    df.rename(columns={
        'Name':                'Frame Property',
        'Number Bars 3-Dir':   '# Long. Bars 3-axis',
        'Number Bars 2-Dir':   '# Long. Bars 2-axis',
        'Clear Cover to Ties': 'Cover',
    }, inplace=True)
    # Corner Bar Size (mm diameter) → Corner Bar Area (mm²)
    if 'Corner Bar Size' in df.columns:
        df['Corner Bar Area'] = df['Corner Bar Size'].map(
            lambda d: round(math.pi * (float(d) / 2) ** 2, 2) if pd.notna(d) else 0.0
        )
    cols = ['Frame Property', '# Long. Bars 3-axis', '# Long. Bars 2-axis',
            'Corner Bar Area', 'Cover']
    return df[[c for c in cols if c in df.columns]]


def _t_beam_rebar(df: pd.DataFrame) -> pd.DataFrame:
    """Frame Sec Def - Conc Beam Reinf → Concrete Beam Rebar Data."""
    df = df.copy()
    df.rename(columns={'Name': 'Frame Property'}, inplace=True)
    # area top = max(Top I-End Area, Top J-End Area)
    for end in ('I', 'J'):
        for side in ('Top', 'Bottom'):
            col = f'{side} {end}-End Area'
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    if 'Top I-End Area' in df.columns and 'Top J-End Area' in df.columns:
        df['area top']    = df[['Top I-End Area',    'Top J-End Area']].max(axis=1)
    if 'Bottom I-End Area' in df.columns and 'Bottom J-End Area' in df.columns:
        df['area bottom'] = df[['Bottom I-End Area', 'Bottom J-End Area']].max(axis=1)
    # E17 engine needs these columns; set #top/#bottom = 0 (not in E23)
    df['#top']    = 0
    df['#bottom'] = 0
    cols = ['Frame Property', '#bottom', 'area bottom', '#top', 'area top', 'Top Cover']
    return df[[c for c in cols if c in df.columns]]


def _t_col_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Conc Col Sum - ACI 318-14 → Concrete Column Summary - ACI 3."""
    df = df.copy()
    df.rename(columns={
        'UniqueName':  'Unique Name',
        'DesignSect':  'Design Section',
        'CornerBarAs': 'Corner Bar As',
        # 'As' se mantiene como 'As' — nombre exacto que espera el motor
        # 'Mid Bar As' ya viene correcto desde E23
    }, inplace=True)
    return df


def _t_beam_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Conc Bm Sum - ACI 318-14 → Concrete Beam Summary - ACI 318."""
    df = df.copy()
    df.rename(columns={
        'UniqueName': 'Unique Name',
        'DesignSect': 'Design Section',
        'AsTop':      'As Top',
        'AsBot':      'As Bottom',
    }, inplace=True)
    return df


def _t_pier_props(df: pd.DataFrame, story_order: list) -> pd.DataFrame:
    """Pier Sec Def - General Pier → Pier Section Properties.

    E23 solo tiene Name/Material; creamos registros por piso con datos vacíos.
    """
    rows = []
    for _, row in df.iterrows():
        pier_name = row.get('Name', '')
        for story in story_order:
            rows.append({'Story': story, 'Pier': pier_name,
                         'Length': 0.0, 'Thickness': 0.2})
    return pd.DataFrame(rows)


def _t_diaphragms(df: pd.DataFrame) -> pd.DataFrame:
    """Joint Assigns - Diaphragms → Joint Assignments - Diaphragms."""
    df = df.copy()
    df.rename(columns={'UniqueName': 'Unique Name'}, inplace=True)
    return df


def _t_local_axes(df: pd.DataFrame, df_summary: pd.DataFrame | None) -> pd.DataFrame:
    """Frame Assigns - Local Axes → Frame Assignments - Local Axes.

    El motor filtra por Design Type == 'Column', así que hay que unir con Frame Assigns - Summary.
    """
    df = df.copy()
    df.rename(columns={'UniqueName': 'Unique Name'}, inplace=True)
    if df_summary is not None:
        type_map = (df_summary[['Story', 'Label', 'Design Type']]
                    .drop_duplicates(subset=['Story', 'Label']))
        df = df.merge(type_map, on=['Story', 'Label'], how='left')
        df['Design Type'] = df['Design Type'].fillna('Beam')
    else:
        df['Design Type'] = 'Beam'
    return df


def _t_offsets(df: pd.DataFrame) -> pd.DataFrame:
    """Frame Assigns - End Len Offsets → Frame Assignments - Offsets."""
    df = df.copy()
    df.rename(columns={
        'UniqueName': 'Unique Name',
        'Offset I':   'Offset I-end',
        'Offset J':   'Offset J-end',
    }, inplace=True)
    return df


# ── Main entry point ──────────────────────────────────────────────────────────

def adapt_e23_to_e17(input_path: str, output_path: str) -> None:
    """
    Lee un XLSX exportado de ETABS 23, transforma todas las hojas al formato
    ETABS 17 y guarda el resultado en output_path.
    """
    xl = pd.ExcelFile(input_path)

    # --- leer todas las hojas necesarias ---
    df_joints_raw    = _read(xl, 'Objects and Elements - Joints')
    df_frames_raw    = _read(xl, 'Objects and Elements - Frames')
    df_areas_raw     = _read(xl, 'Objects and Elements - Areas')
    df_floor_conn    = _read(xl, 'Floor Object Connectivity')
    df_conc_data     = _read(xl, 'Mat Prop - Concrete Data')
    df_mech_props    = _read(xl, 'Mat Prop - Basic Mech Props')
    df_mass_diaphr   = _read(xl, 'Mass Summary by Diaphragm')
    df_restraints    = _read(xl, 'Joint Assigns - Restraints')
    df_fr_prop_sum   = _read(xl, 'Frame Prop - Summary')
    df_fr_sect_prop  = _read(xl, 'Frame Assigns - Sect Prop')
    df_fr_summary    = _read(xl, 'Frame Assigns - Summary')
    df_area_sect_prop= _read(xl, 'Area Assigns - Sect Prop')
    df_slab_prop     = _read(xl, 'Slab Property Definitions')
    df_wall_prop     = _read(xl, 'Wall Property Def - Specified')
    df_area_loads    = _read(xl, 'Area Loads - Uniform')
    df_frame_loads   = _read(xl, 'Frame Loads - Distributed')
    df_col_reinf     = _read(xl, 'Frame Sec Def - Conc Col Reinf')
    df_beam_reinf    = _read(xl, 'Frame Sec Def - Conc Beam Reinf')
    df_col_sum       = _read(xl, 'Conc Col Sum - ACI 318-14')
    df_beam_sum      = _read(xl, 'Conc Bm Sum - ACI 318-14')
    df_pier_defs     = _read(xl, 'Pier Sec Def - General Pier')
    df_diaphr_asgns  = _read(xl, 'Joint Assigns - Diaphragms')
    df_local_axes    = _read(xl, 'Frame Assigns - Local Axes')
    df_offsets       = _read(xl, 'Frame Assigns - End Len Offsets')

    # Derivar story_order para pier props (bottom → top)
    story_order = []
    if df_mass_diaphr is not None and 'Story' in df_mass_diaphr.columns:
        story_order = df_mass_diaphr['Story'].tolist()

    # --- transformar ---
    df_j  = _t_joints(df_joints_raw) if df_joints_raw is not None else None
    df_f  = _t_frames(df_frames_raw) if df_frames_raw is not None else None
    df_sh = _t_shells(df_areas_raw)  if df_areas_raw  is not None else None
    df_fc = _t_floor_conn(df_floor_conn) if df_floor_conn is not None else None
    df_m  = _t_materials(df_conc_data, df_mech_props) if df_conc_data is not None else None
    df_r  = _t_restraints(df_restraints) if df_restraints is not None else None
    df_fs = _t_frame_sections(df_fr_prop_sum) if df_fr_prop_sum is not None else None
    df_fa = _t_frame_assigns(df_fr_sect_prop, df_fr_summary) if df_fr_sect_prop is not None else None
    df_sa = _t_shell_assigns(df_area_sect_prop) if df_area_sect_prop is not None else None
    df_sl = _t_slab_sections(df_slab_prop) if df_slab_prop is not None else None
    df_wl = _t_wall_sections(df_wall_prop) if df_wall_prop is not None else None
    df_al = _t_shell_loads(df_area_loads) if df_area_loads is not None else None
    df_fl = _t_frame_loads(df_frame_loads) if df_frame_loads is not None else None
    df_cr = _t_col_rebar(df_col_reinf) if df_col_reinf is not None else None
    df_br = _t_beam_rebar(df_beam_reinf) if df_beam_reinf is not None else None
    df_cs = _t_col_summary(df_col_sum) if df_col_sum is not None else None
    df_bs = _t_beam_summary(df_beam_sum) if df_beam_sum is not None else None
    df_pp = _t_pier_props(df_pier_defs, story_order) if df_pier_defs is not None else None
    df_da = _t_diaphragms(df_diaphr_asgns) if df_diaphr_asgns is not None else None
    df_la = _t_local_axes(df_local_axes, df_fr_summary) if df_local_axes is not None else None
    df_of = _t_offsets(df_offsets) if df_offsets is not None else None

    # --- escribir XLSX E17-compatible ---
    wb = Workbook()
    wb.remove(wb.active)

    if df_j is not None:
        _write_sheet(wb, 'Objects and Elements - Joints',
                     'TABLE:  "OBJECTS AND ELEMENTS - JOINTS"',
                     ['Story', 'Element Label', 'Object Type', 'Object Label',
                      'Global X', 'Global Y', 'Global Z'],
                     [None, None, None, None, 'm', 'm', 'm'],
                     df_j)

    if df_f is not None:
        _write_sheet(wb, 'Objects and Elements - Frames',
                     'TABLE:  "OBJECTS AND ELEMENTS - FRAMES"',
                     ['Story', 'Element Label', 'Object Type', 'Object Label',
                      'Joint I', 'Joint J'],
                     [None, None, None, None, None, None],
                     df_f)

    if df_sh is not None:
        _write_sheet(wb, 'Objects and Elements - Shells',
                     'TABLE:  "OBJECTS AND ELEMENTS - SHELLS"',
                     ['Story', 'Element Label', 'Area Type', 'Area Label',
                      'Joint 1', 'Joint 2', 'Joint 3', 'Joint 4'],
                     [None, None, None, None, None, None, None, None],
                     df_sh)

    if df_fc is not None:
        _write_sheet(wb, 'Floor Connectivity',
                     'TABLE:  "FLOOR CONNECTIVITY"',
                     ['Story', 'Unique Name', 'Points', 'Perimeter', 'Area'],
                     [None, None, None, 'm', 'm2'],
                     df_fc)

    if df_m is not None:
        _write_sheet(wb, 'Material Properties - Concrete',
                     'TABLE:  "MATERIAL PROPERTIES - CONCRETE"',
                     ['Name', 'Fc', 'E', 'G', 'Poisson', 'Lightweight?'],
                     [None, 'MPa', 'MPa', 'MPa', None, None],
                     df_m)

    if df_mass_diaphr is not None:
        dm = df_mass_diaphr.copy()
        # E17 exporta masa en kg y momento de inercia en ton·m² — convertir desde ton/ton·m²
        for col in ['Mass X', 'Mass Y']:
            if col in dm.columns:
                dm[col] = pd.to_numeric(dm[col], errors='coerce') * 1000.0  # ton → kg
        if 'Mass Moment of Inertia' in dm.columns:
            dm['Mass Moment of Inertia'] = pd.to_numeric(
                dm['Mass Moment of Inertia'], errors='coerce')  # ton·m² ya está correcto
        _write_sheet(wb, 'Mass Summary by Diaphragm',
                     'TABLE:  "MASS SUMMARY BY DIAPHRAGM"',
                     ['Story', 'Diaphragm', 'Mass X', 'Mass Y',
                      'Mass Moment of Inertia', 'X Mass Center', 'Y Mass Center'],
                     [None, None, 'kg', 'kg', 'ton-m2', 'm', 'm'],
                     dm)

    if df_r is not None:
        _write_sheet(wb, 'Joint Assignments - Restraints',
                     'TABLE:  "JOINT ASSIGNMENTS - RESTRAINTS"',
                     ['Story', 'Unique Name', 'UX', 'UY', 'UZ', 'RX', 'RY', 'RZ'],
                     [None, None, None, None, None, None, None, None],
                     df_r)

    if df_fs is not None:
        _write_sheet(wb, 'Frame Sections',
                     'TABLE:  "FRAME SECTIONS"',
                     ['Name', 'Material', 'Shape', 't3', 't2',
                      'Area', 'I33', 'I22', 'J'],
                     [None, None, None, 'm', 'm', 'm2', 'm4', 'm4', 'm4'],
                     df_fs)

    if df_fa is not None:
        _write_sheet(wb, 'Frame Assignments - Sections',
                     'TABLE:  "FRAME ASSIGNMENTS - SECTIONS"',
                     ['Story', 'Label', 'Unique Name', 'Analysis Section',
                      'Design Section', 'Design Type', 'Section Type'],
                     [None, None, None, None, None, None, None],
                     df_fa)

    if df_sl is not None:
        _write_sheet(wb, 'Shell Sections - Slab',
                     'TABLE:  "SHELL SECTIONS - SLAB"',
                     ['Name', 'Material', 'Slab Thickness',
                      'One-Way Load Distribution?'],
                     [None, None, 'm', None],
                     df_sl)

    if df_wl is not None:
        _write_sheet(wb, 'Shell Sections - Wall',
                     'TABLE:  "SHELL SECTIONS - WALL"',
                     ['Name', 'Material', 'Thickness'],
                     [None, None, 'm'],
                     df_wl)

    if df_sa is not None:
        _write_sheet(wb, 'Shell Assignments - Sections',
                     'TABLE:  "SHELL ASSIGNMENTS - SECTIONS"',
                     ['Story', 'Unique Name', 'Section'],
                     [None, None, None],
                     df_sa)

    if df_fl is not None:
        fl_cols = ['Story', 'Label', 'Unique Name', 'Load Pattern',
                   'Direction', 'Distance A', 'Distance B', 'Load A', 'Load B']
        if 'Force at End' in df_fl.columns:
            fl_cols.append('Force at End')
        _write_sheet(wb, 'Frame Loads - Distributed',
                     'TABLE:  "FRAME LOADS - DISTRIBUTED"',
                     fl_cols,
                     [None] * len(fl_cols),
                     df_fl)

    if df_al is not None:
        _write_sheet(wb, 'Shell Loads - Uniform',
                     'TABLE:  "SHELL LOADS - UNIFORM"',
                     ['Story', 'Label', 'Unique Name', 'Load Pattern',
                      'Direction', 'Load'],
                     [None, None, None, None, None, 'kN/m2'],
                     df_al)

    if df_cr is not None:
        _write_sheet(wb, 'Concrete Column Rebar Data',
                     'TABLE:  "CONCRETE COLUMN REBAR DATA"',
                     ['Frame Property', '# Long. Bars 3-axis',
                      '# Long. Bars 2-axis', 'Corner Bar Area', 'Cover'],
                     [None, None, None, 'mm2', 'm'],
                     df_cr)

    if df_br is not None:
        _write_sheet(wb, 'Concrete Beam Rebar Data',
                     'TABLE:  "CONCRETE BEAM REBAR DATA"',
                     ['Frame Property', '#bottom', 'area bottom',
                      '#top', 'area top', 'Top Cover'],
                     [None, None, 'm2', None, 'm2', 'm'],
                     df_br)

    if df_cs is not None:
        _write_sheet(wb, 'Concrete Column Summary - ACI 3',
                     'TABLE:  "CONCRETE COLUMN SUMMARY - ACI3"',
                     list(df_cs.columns),
                     [None] * len(df_cs.columns),
                     df_cs)

    if df_bs is not None:
        _write_sheet(wb, 'Concrete Beam Summary - ACI 318',
                     'TABLE:  "CONCRETE BEAM SUMMARY - ACI3"',
                     list(df_bs.columns),
                     [None] * len(df_bs.columns),
                     df_bs)

    if df_pp is not None and len(df_pp):
        _write_sheet(wb, 'Pier Section Properties',
                     'TABLE:  "PIER SECTION PROPERTIES"',
                     ['Story', 'Pier', 'Length', 'Thickness'],
                     [None, None, 'm', 'm'],
                     df_pp)

    if df_da is not None:
        _write_sheet(wb, 'Joint Assignments - Diaphragms',
                     'TABLE:  "JOINT ASSIGNMENTS - DIAPHRAGMS"',
                     list(df_da.columns),
                     [None] * len(df_da.columns),
                     df_da)

    if df_la is not None:
        _write_sheet(wb, 'Frame Assignments - Local Axes',
                     'TABLE:  "FRAME ASSIGNMENTS - LOCAL AXES"',
                     list(df_la.columns),
                     [None] * len(df_la.columns),
                     df_la)

    if df_of is not None:
        _write_sheet(wb, 'Frame Assignments - Offsets',
                     'TABLE:  "FRAME ASSIGNMENTS - OFFSETS"',
                     list(df_of.columns),
                     [None] * len(df_of.columns),
                     df_of)

    wb.save(output_path)
