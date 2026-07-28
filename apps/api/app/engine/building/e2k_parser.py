"""
E2K Parser: Converts ETABS .e2k text export → input_model.xlsx

Sheet row convention expected by the engine:
  Row 1: TABLE header string
  Row 2: column names
  Row 3: units row
  Row 4+: data
"""
import re
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ── helpers ───────────────────────────────────────────────────────────────────

def _tok(line: str) -> list:
    """Tokenise: quoted strings become one token, bare words another."""
    return [m.group(1) if m.group(1) is not None else m.group(2)
            for m in re.finditer(r'"([^"]*)"|([\S]+)', line)]


def _kv_from(tokens: list, start: int) -> dict:
    """Pair tokens[start], tokens[start+1], tokens[start+2], … as key→value."""
    d = {}
    i = start
    while i + 1 < len(tokens):
        d[tokens[i].upper()] = tokens[i + 1]
        i += 2
    return d


def _write_sheet(wb: Workbook, sheet_name: str, table_key: str,
                 columns: list, units: list, df: pd.DataFrame) -> None:
    """Write one sheet with the 3-row header the engine expects."""
    ws = wb.create_sheet(title=sheet_name)
    ws.append([table_key])
    ws.append(columns)
    ws.append(units)
    if df is not None and len(df):
        for row in dataframe_to_rows(df[columns], index=False, header=False):
            ws.append(row)


# ── parser class ──────────────────────────────────────────────────────────────

class E2KParser:

    def __init__(self, e2k_path: str):
        self.e2k_path = str(e2k_path)

        self._sections: dict[str, list] = {}

        self.stories: dict[str, float] = {}    # story → Z elevation (m)
        self.story_order: list = []            # bottom → top

        self.mat_concrete: dict = {}           # mat_name → {Fc, Ec, G, Poisson}
        self.frame_sec_defs: dict = {}         # sec_name → {Material, t3, t2, …}
        self.concrete_secs: dict = {}          # sec_name → {type, bars_3, bars_2, area, cover}
        self.slab_props: dict = {}             # prop_name → {Material, Thickness, OneWay}
        self.wall_props: dict = {}             # prop_name → {Material, Thickness}

        self.points: dict = {}                 # pt_name → (X, Y)
        self.point_assigns: dict = {}          # (pt, story) → True

        self.line_conns: dict = {}             # line_name → {type, pt_i, pt_j}
        self.line_assigns: dict = {}           # (line, story) → {section}
        self.area_conns: dict = {}             # area_name → {type, pts}
        self.area_assigns: dict = {}           # (area, story) → {section}

        self.frame_loads: list = []
        self.shell_loads: list = []

        self.load_patterns: dict = {}          # pat_name → type_str

        self.mass_source_loads: list = []      # [{lc, factor}]
        self.mass_include_elements: bool = True

        self.pier_names: list = []

        # element labels — assigned after parsing
        self._joint_label: dict = {}           # (pt, story) → int
        self._frame_elem_label: dict = {}      # (line, story) → int
        self._shell_elem_label: dict = {}      # (area, story) → int

    # ── public ────────────────────────────────────────────────────────────────

    def parse(self) -> None:
        self._read_sections()
        self._parse_stories()
        self._parse_load_patterns()
        self._parse_materials()
        self._parse_frame_sections()
        self._parse_concrete_sections()
        self._parse_shell_props()        # handles both SLAB and WALL sections
        self._parse_points()
        self._parse_point_assigns()
        self._parse_line_conns()
        self._parse_line_assigns()
        self._parse_area_conns()
        self._parse_area_assigns()
        self._parse_frame_loads()
        self._parse_shell_loads()
        self._parse_mass_source()
        self._parse_pier_spandrel_names()
        self._assign_element_labels()

    def write_xlsx(self, output_path: str,
                   supplemental_xlsx_path: str | None = None) -> None:
        df_joints     = self._build_oe_joints()
        df_frames     = self._build_oe_frames()
        df_shells     = self._build_oe_shells()
        df_floor_conn = self._build_floor_connectivity(df_shells, df_joints)
        df_mats       = self._build_materials()
        df_mass       = self._build_mass_summary(df_joints, df_shells)
        df_restraints = self._build_restraints(df_joints)
        df_fr_sec     = self._build_frame_sections()
        df_fr_asgn    = self._build_frame_assignments(df_frames)
        df_sh_sec_slab = self._build_shell_sections_slab()
        df_sh_sec_wall = self._build_shell_sections_wall()
        df_sh_asgn    = self._build_shell_assignments(df_shells)
        df_fr_loads   = self._build_frame_loads_df()
        df_sh_loads   = self._build_shell_loads_df()
        df_col_rebar  = self._build_column_rebar_data()

        wb = Workbook()
        wb.remove(wb.active)

        _write_sheet(wb, 'Objects and Elements - Joints',
                     'TABLE:  "OBJECTS AND ELEMENTS - JOINTS"',
                     ['Story', 'Element Label', 'Object Type', 'Object Label',
                      'Global X', 'Global Y', 'Global Z'],
                     [None, None, None, None, 'm', 'm', 'm'],
                     df_joints)

        _write_sheet(wb, 'Objects and Elements - Frames',
                     'TABLE:  "OBJECTS AND ELEMENTS - FRAMES"',
                     ['Story', 'Element Label', 'Object Type', 'Object Label',
                      'Joint I', 'Joint J'],
                     [None, None, None, None, None, None],
                     df_frames)

        _write_sheet(wb, 'Objects and Elements - Shells',
                     'TABLE:  "OBJECTS AND ELEMENTS - SHELLS"',
                     ['Story', 'Element Label', 'Area Type', 'Area Label',
                      'Joint 1', 'Joint 2', 'Joint 3', 'Joint 4'],
                     [None, None, None, None, None, None, None, None],
                     df_shells)

        _write_sheet(wb, 'Floor Connectivity',
                     'TABLE:  "FLOOR CONNECTIVITY"',
                     ['Story', 'Unique Name', 'Points', 'Perimeter', 'Area'],
                     [None, None, None, 'm', 'm2'],
                     df_floor_conn)

        _write_sheet(wb, 'Material Properties - Concrete',
                     'TABLE:  "MATERIAL PROPERTIES - CONCRETE"',
                     ['Name', 'Fc', 'Ec', 'G', 'Poisson', 'Lightweight?'],
                     [None, 'MPa', 'MPa', 'MPa', None, None],
                     df_mats)

        _write_sheet(wb, 'Mass Summary by Diaphragm',
                     'TABLE:  "MASS SUMMARY BY DIAPHRAGM"',
                     ['Story', 'Diaphragm', 'Mass X', 'Mass Y',
                      'Mass Moment of Inertia', 'X Mass Center', 'Y Mass Center'],
                     [None, None, 'ton', 'ton', 'ton-m2', 'm', 'm'],
                     df_mass)

        _write_sheet(wb, 'Joint Assignments - Restraints',
                     'TABLE:  "JOINT ASSIGNMENTS - RESTRAINTS"',
                     ['Story', 'Unique Name', 'UX', 'UY', 'UZ', 'RX', 'RY', 'RZ'],
                     [None, None, None, None, None, None, None, None],
                     df_restraints)

        _write_sheet(wb, 'Frame Sections',
                     'TABLE:  "FRAME SECTIONS"',
                     ['Name', 'Material', 'Shape', 't3', 't2',
                      'Area', 'I33', 'I22', 'J'],
                     [None, None, None, 'm', 'm', 'm2', 'm4', 'm4', 'm4'],
                     df_fr_sec)

        _write_sheet(wb, 'Frame Assignments - Sections',
                     'TABLE:  "FRAME ASSIGNMENTS - SECTIONS"',
                     ['Story', 'Label', 'Unique Name', 'Analysis Section',
                      'Design Section'],
                     [None, None, None, None, None],
                     df_fr_asgn)

        _write_sheet(wb, 'Shell Sections - Slab',
                     'TABLE:  "SHELL SECTIONS - SLAB"',
                     ['Name', 'Material', 'Slab Thickness',
                      'One-Way Load Distribution?'],
                     [None, None, 'm', None],
                     df_sh_sec_slab)

        if df_sh_sec_wall is not None and len(df_sh_sec_wall):
            _write_sheet(wb, 'Shell Sections - Wall',
                         'TABLE:  "SHELL SECTIONS - WALL"',
                         ['Name', 'Material', 'Thickness'],
                         [None, None, 'm'],
                         df_sh_sec_wall)

        _write_sheet(wb, 'Shell Assignments - Sections',
                     'TABLE:  "SHELL ASSIGNMENTS - SECTIONS"',
                     ['Story', 'Unique Name', 'Section'],
                     [None, None, None],
                     df_sh_asgn)

        _write_sheet(wb, 'Frame Loads - Distributed',
                     'TABLE:  "FRAME LOADS - DISTRIBUTED"',
                     ['Story', 'Label', 'Unique Name', 'Load Pattern',
                      'Direction', 'Distance A', 'Distance B',
                      'Load A', 'Load B'],
                     [None, None, None, None, None, None, None, 'kN/m', 'kN/m'],
                     df_fr_loads)

        _write_sheet(wb, 'Shell Loads - Uniform',
                     'TABLE:  "SHELL LOADS - UNIFORM"',
                     ['Story', 'Label', 'Unique Name', 'Load Pattern',
                      'Direction', 'Load'],
                     [None, None, None, None, None, 'kN/m2'],
                     df_sh_loads)

        if df_col_rebar is not None and len(df_col_rebar):
            _write_sheet(wb, 'Concrete Column Rebar Data',
                         'TABLE:  "CONCRETE COLUMN REBAR DATA"',
                         ['Frame Property', '# Long. Bars 3-axis',
                          '# Long. Bars 2-axis', 'Corner Bar Area', 'Cover'],
                         [None, None, None, 'mm2', 'm'],
                         df_col_rebar)

        if self.pier_names:
            _write_sheet(wb, 'Pier Section Properties',
                         'TABLE:  "PIER SECTION PROPERTIES"',
                         ['Story', 'Pier', 'Length', 'Thickness'],
                         [None, None, 'm', 'm'],
                         self._build_pier_section_properties())

            _write_sheet(wb, 'Shell Assignments - Pier Spandr',
                         'TABLE:  "SHELL ASSIGNMENTS - PIER SPANDR"',
                         ['Story', 'Unique Name', 'Pier'],
                         [None, None, None],
                         self._build_shell_pier_assignments(df_shells))

        if supplemental_xlsx_path:
            self._merge_supplemental_sheets(wb, supplemental_xlsx_path)

        wb.save(output_path)

    # ── private: file reading ─────────────────────────────────────────────────

    def _read_sections(self) -> None:
        current = None
        buf: list = []
        with open(self.e2k_path, 'r', encoding='latin-1') as fh:
            for raw in fh:
                stripped = raw.strip()
                if stripped.startswith('$ ') and not stripped.upper().startswith('$ END'):
                    if current is not None:
                        self._sections[current] = buf
                    current = stripped[2:].strip().upper()
                    buf = []
                elif stripped.upper().startswith('$ END'):
                    if current is not None:
                        self._sections[current] = buf
                    current = None
                    buf = []
                elif current is not None and stripped:
                    buf.append(stripped)
        if current is not None:
            self._sections[current] = buf

    # ── per-section parsers ───────────────────────────────────────────────────

    def _parse_stories(self) -> None:
        lines = self._sections.get('STORIES - IN SEQUENCE FROM TOP', [])
        raw: list = []
        for line in lines:
            t = _tok(line)
            if not t or t[0].upper() != 'STORY':
                continue
            name = t[1]
            kv = _kv_from(t, 2)
            if 'ELEV' in kv:
                raw.append((name, float(kv['ELEV']), True))
            elif 'HEIGHT' in kv:
                raw.append((name, float(kv['HEIGHT']), False))

        if not raw:
            return
        raw.reverse()  # file is top→bottom; we want bottom→top

        z = 0.0
        for name, val, is_base in raw:
            z = val if is_base else z + val
            self.stories[name] = z
        self.story_order = [s[0] for s in raw]

    def _parse_load_patterns(self) -> None:
        for line in self._sections.get('LOAD PATTERNS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'LOADPATTERN':
                continue
            name = t[1]
            kv = _kv_from(t, 2)
            self.load_patterns[name] = kv.get('TYPE', 'Other')

    def _parse_materials(self) -> None:
        mat_data: dict = {}
        concrete_names: set = set()
        for line in self._sections.get('MATERIAL PROPERTIES', []):
            t = _tok(line)
            if not t or t[0].upper() != 'MATERIAL':
                continue
            name = t[1]
            if name not in mat_data:
                mat_data[name] = {}
            kv = _kv_from(t, 2)
            if kv.get('TYPE', '').upper() == 'CONCRETE':
                concrete_names.add(name)
            mat_data[name].update(kv)

        for name in concrete_names:
            kv = mat_data[name]
            fc_kpa  = abs(float(kv.get('FC', 0)))
            fc_mpa  = fc_kpa / 1000.0
            e_kpa   = abs(float(kv.get('E', 0)))
            e_mpa   = e_kpa / 1000.0
            nu      = float(kv.get('U', 0.2))
            g_mpa   = e_mpa / (2 * (1 + nu)) if e_mpa > 0 else 0.0
            self.mat_concrete[name] = {
                'Fc': round(fc_mpa, 4),
                'Ec': round(e_mpa, 4),
                'G': round(g_mpa, 4),
                'Poisson': nu,
            }

    def _parse_frame_sections(self) -> None:
        # Multi-line: first line has MATERIAL/SHAPE/D/B, second may have JMOD etc.
        sec_data: dict = {}
        for line in self._sections.get('FRAME SECTIONS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'FRAMESECTION':
                continue
            name = t[1]
            if name not in sec_data:
                sec_data[name] = {}
            kv = _kv_from(t, 2)
            sec_data[name].update(kv)

        for name, kv in sec_data.items():
            mat = kv.get('MATERIAL', '')
            shape = kv.get('SHAPE', 'Rectangular')
            t3 = float(kv.get('D', kv.get('T3', 0)))
            t2 = float(kv.get('B', kv.get('T2', t3)))
            area = t3 * t2
            i33  = t2 * t3 ** 3 / 12.0
            i22  = t3 * t2 ** 3 / 12.0
            j    = i33 + i22
            self.frame_sec_defs[name] = {
                'Material': mat,
                'Shape': shape,
                't3': t3,
                't2': t2,
                'Area': area,
                'I33': i33,
                'I22': i22,
                'J': j,
            }

    def _parse_concrete_sections(self) -> None:
        for line in self._sections.get('CONCRETE SECTIONS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'CONCRETESECTION':
                continue
            name = t[1]
            kv = _kv_from(t, 2)
            ctype   = kv.get('TYPE', '').capitalize()
            pattern = kv.get('PATTERN', '')
            longbararea = float(kv.get('LONGBARAREA', 0))
            cover   = float(kv.get('COVER', kv.get('COVERTOP', 0.04)))

            bars_3 = bars_2 = 0
            m = re.match(r'R-(\d+)-(\d+)', pattern)
            if m:
                bars_3 = int(m.group(1))
                bars_2 = int(m.group(2))

            self.concrete_secs[name] = {
                'type': ctype,
                'bars_3axis': bars_3,
                'bars_2axis': bars_2,
                'longbararea_m2': longbararea,
                'cover': cover,
            }

    def _parse_shell_props(self) -> None:
        """Parses both SLAB PROPERTIES and WALL PROPERTIES sections (both use SHELLPROP)."""
        for sec_name in ('SLAB PROPERTIES', 'WALL PROPERTIES'):
            prop_data: dict = {}
            for line in self._sections.get(sec_name, []):
                t = _tok(line)
                if not t or t[0].upper() != 'SHELLPROP':
                    continue
                name = t[1]
                if name not in prop_data:
                    prop_data[name] = {}
                kv = _kv_from(t, 2)
                prop_data[name].update(kv)

            for name, kv in prop_data.items():
                proptype = kv.get('PROPTYPE', 'Slab').strip('"').capitalize()
                mat = kv.get('MATERIAL', '')
                if proptype.lower() == 'slab':
                    thick = float(kv.get('SLABTHICKNESS', kv.get('MEMBTHICKNESS', 0.2)))
                    one_way = kv.get('ONEWAYLOADDIST', 'No').strip('"')
                    self.slab_props[name] = {
                        'Material': mat,
                        'Thickness': thick,
                        'OneWay': one_way,
                    }
                elif proptype.lower() == 'wall':
                    thick = float(kv.get('WALLTHICKNESS', 0.2))
                    self.wall_props[name] = {'Material': mat, 'Thickness': thick}

    def _parse_points(self) -> None:
        for line in self._sections.get('POINT COORDINATES', []):
            t = _tok(line)
            if not t or t[0].upper() != 'POINT':
                continue
            try:
                self.points[t[1]] = (float(t[2]), float(t[3]))
            except (IndexError, ValueError):
                pass

    def _parse_point_assigns(self) -> None:
        for line in self._sections.get('POINT ASSIGNS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'POINTASSIGN':
                continue
            if len(t) < 3:
                continue
            self.point_assigns[(t[1], t[2])] = True

    def _parse_line_conns(self) -> None:
        for line in self._sections.get('LINE CONNECTIVITIES', []):
            t = _tok(line)
            if not t or t[0].upper() != 'LINE':
                continue
            if len(t) < 5:
                continue
            # LINE "name" COLUMN/BEAM "pt_i" "pt_j" ...
            self.line_conns[t[1]] = {
                'type': t[2].upper(),
                'pt_i': t[3],
                'pt_j': t[4],
            }

    def _parse_line_assigns(self) -> None:
        for line in self._sections.get('LINE ASSIGNS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'LINEASSIGN':
                continue
            if len(t) < 3:
                continue
            name, story = t[1], t[2]
            kv = _kv_from(t, 3)
            self.line_assigns[(name, story)] = {'section': kv.get('SECTION', '')}

    def _parse_area_conns(self) -> None:
        for line in self._sections.get('AREA CONNECTIVITIES', []):
            t = _tok(line)
            if not t or t[0].upper() != 'AREA':
                continue
            if len(t) < 4:
                continue
            # AREA "name" FLOOR/WALL num_pts "pt1" "pt2" ...
            name = t[1]
            atype = t[2].upper()   # FLOOR or WALL
            try:
                n = int(t[3])
            except ValueError:
                continue
            pts = t[4: 4 + n]
            self.area_conns[name] = {'type': atype, 'pts': pts}

    def _parse_area_assigns(self) -> None:
        for line in self._sections.get('AREA ASSIGNS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'AREAASSIGN':
                continue
            if len(t) < 3:
                continue
            name, story = t[1], t[2]
            kv = _kv_from(t, 3)
            self.area_assigns[(name, story)] = {'section': kv.get('SECTION', '')}

    def _parse_frame_loads(self) -> None:
        for line in self._sections.get('FRAME OBJECT LOADS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'LINELOAD':
                continue
            if len(t) < 3:
                continue
            frame_name, story = t[1], t[2]
            kv = _kv_from(t, 3)
            self.frame_loads.append({
                'frame': frame_name,
                'story': story,
                'lc':        kv.get('LC', ''),
                'direction': kv.get('DIR', 'Z'),
                'fval':      float(kv.get('FVAL', 0)),
            })

    def _parse_shell_loads(self) -> None:
        for line in self._sections.get('SHELL OBJECT LOADS', []):
            t = _tok(line)
            if not t or t[0].upper() != 'AREALOAD':
                continue
            if len(t) < 3:
                continue
            area_name, story = t[1], t[2]
            kv = _kv_from(t, 3)
            self.shell_loads.append({
                'area':      area_name,
                'story':     story,
                'lc':        kv.get('LC', ''),
                'direction': kv.get('DIR', 'Z'),
                'load':      float(kv.get('FVAL', 0)),
            })

    def _parse_mass_source(self) -> None:
        for line in self._sections.get('MASS SOURCE', []):
            t = _tok(line)
            if not t:
                continue
            key = t[0].upper()
            if key == 'MASSSOURCE':
                kv = _kv_from(t, 2)
                inc = kv.get('INCLUDEELEMENTS', 'Yes')
                self.mass_include_elements = inc.strip('"').upper() == 'YES'
            elif key == 'MASSSOURCELOAD':
                # MASSSOURCELOAD "source_name" "lc_name" factor
                if len(t) >= 3:
                    lc     = t[2]
                    factor = float(t[3]) if len(t) > 3 else 1.0
                    self.mass_source_loads.append({'lc': lc, 'factor': factor})

    def _parse_pier_spandrel_names(self) -> None:
        for line in self._sections.get('PIER/SPANDREL NAMES', []):
            t = _tok(line)
            if t and t[0].upper() == 'PIERNAME' and len(t) > 1:
                self.pier_names.append(t[1])

    # ── element label assignment ──────────────────────────────────────────────

    def _assign_element_labels(self) -> None:
        counter = 1
        for key in sorted(self.point_assigns.keys()):
            self._joint_label[key] = counter
            counter += 1
        for key in sorted(self.line_assigns.keys()):
            self._frame_elem_label[key] = counter
            counter += 1
        for key in sorted(self.area_assigns.keys()):
            self._shell_elem_label[key] = counter
            counter += 1

    def _story_below(self, story: str):
        idx = self.story_order.index(story)
        return self.story_order[idx - 1] if idx > 0 else None

    # ── DataFrame builders ────────────────────────────────────────────────────

    def _build_oe_joints(self) -> pd.DataFrame:
        rows = []
        for (pt, story), label in self._joint_label.items():
            x, y = self.points.get(pt, (0.0, 0.0))
            z = self.stories.get(story, 0.0)
            rows.append({
                'Story': story,
                'Element Label': label,
                'Object Type': 'Joint',
                'Object Label': pt,
                'Global X': x,
                'Global Y': y,
                'Global Z': z,
            })
        return pd.DataFrame(rows)

    def _build_oe_frames(self) -> pd.DataFrame:
        rows = []
        for (ln, story), label in self._frame_elem_label.items():
            conn = self.line_conns.get(ln, {})
            ltype = conn.get('type', 'BEAM')
            pt_i  = conn.get('pt_i', '')
            pt_j  = conn.get('pt_j', '')

            if ltype == 'COLUMN':
                s_below = self._story_below(story)
                ji = self._joint_label.get((pt_i, s_below), 0)
                jj = self._joint_label.get((pt_i, story), 0)
                obj_type = 'Column'
            else:
                ji = self._joint_label.get((pt_i, story), 0)
                jj = self._joint_label.get((pt_j, story), 0)
                obj_type = 'Frame'

            rows.append({
                'Story': story,
                'Element Label': label,
                'Object Type': obj_type,
                'Object Label': ln,
                'Joint I': ji,
                'Joint J': jj,
            })
        return pd.DataFrame(rows)

    def _build_oe_shells(self) -> pd.DataFrame:
        rows = []
        for (ar, story), label in self._shell_elem_label.items():
            conn  = self.area_conns.get(ar, {})
            pts   = conn.get('pts', [])
            atype = conn.get('type', 'FLOOR').capitalize()

            js = [self._joint_label.get((p, story), 0) for p in pts[:4]]
            while len(js) < 4:
                js.append(js[-1] if js else 0)

            rows.append({
                'Story': story,
                'Element Label': label,
                'Area Type': atype,
                'Area Label': ar,
                'Joint 1': js[0],
                'Joint 2': js[1],
                'Joint 3': js[2],
                'Joint 4': js[3],
            })
        return pd.DataFrame(rows)

    def _build_floor_connectivity(self, df_shells: pd.DataFrame,
                                  df_joints: pd.DataFrame) -> pd.DataFrame:
        jcoords = {
            row['Element Label']: (row['Global X'], row['Global Y'])
            for _, row in df_joints.iterrows()
        }
        rows = []
        for _, row in df_shells.iterrows():
            if str(row.get('Area Type', '')).lower() != 'floor':
                continue
            seen = []
            for i in range(1, 5):
                j = row.get(f'Joint {i}', 0)
                if j and j not in seen:
                    seen.append(j)
            coords = [jcoords.get(j, (0.0, 0.0)) for j in seen]
            n = len(coords)
            perim = area = 0.0
            for i in range(n):
                x1, y1 = coords[i]
                x2, y2 = coords[(i + 1) % n]
                perim += math.hypot(x2 - x1, y2 - y1)
                area  += x1 * y2 - x2 * y1
            area = abs(area) / 2.0
            rows.append({
                'Story':       row['Story'],
                'Unique Name': row['Element Label'],
                'Points':      ';'.join(str(int(j)) for j in seen),
                'Perimeter':   round(perim, 6),
                'Area':        round(area, 6),
            })
        return pd.DataFrame(rows)

    def _build_materials(self) -> pd.DataFrame:
        rows = [
            {'Name': n, 'Fc': m['Fc'], 'Ec': m['Ec'], 'G': m['G'],
             'Poisson': m['Poisson'], 'Lightweight?': 'No'}
            for n, m in self.mat_concrete.items()
        ]
        return pd.DataFrame(rows)

    def _build_mass_summary(self, df_joints: pd.DataFrame,
                             df_shells: pd.DataFrame) -> pd.DataFrame:
        G_CONCRETE = 24.0
        GRAVITY    = 9.81

        jcoords = {
            row['Element Label']: (row['Global X'], row['Global Y'])
            for _, row in df_joints.iterrows()
        }

        shell_area: dict  = {}
        shell_centroid: dict = {}
        for _, row in df_shells.iterrows():
            if str(row.get('Area Type', '')).lower() != 'floor':
                continue
            seen = []
            for i in range(1, 5):
                j = row.get(f'Joint {i}', 0)
                if j and j not in seen:
                    seen.append(j)
            coords = [jcoords.get(j, (0.0, 0.0)) for j in seen]
            n = len(coords)
            area = cx = cy = 0.0
            for i in range(n):
                x1, y1 = coords[i]
                x2, y2 = coords[(i + 1) % n]
                cross = x1 * y2 - x2 * y1
                area += cross
                cx   += (x1 + x2) * cross
                cy   += (y1 + y2) * cross
            area = abs(area) / 2.0
            if area > 1e-9:
                cx = abs(cx) / (6 * area)
                cy = abs(cy) / (6 * area)
            key = (row['Area Label'], row['Story'])
            shell_area[key]     = area
            shell_centroid[key] = (cx, cy)

        mass_lc_names = {e['lc'] for e in self.mass_source_loads}

        rows = []
        for i, story in enumerate(self.story_order[1:], start=1):
            total_mass = cx_sum = cy_sum = mmi = 0.0
            for (area_label, ar_story), area in shell_area.items():
                if ar_story != story:
                    continue
                asgn = self.area_assigns.get((area_label, story), {})
                sec  = asgn.get('section', '')
                sp   = self.slab_props.get(sec, {})
                thick = sp.get('Thickness', 0.2)

                m_sw = area * thick * G_CONCRETE / GRAVITY if self.mass_include_elements else 0.0
                m_imp = 0.0
                for sl in self.shell_loads:
                    if sl['area'] == area_label and sl['story'] == story \
                            and sl['lc'] in mass_lc_names:
                        factor = next((e['factor'] for e in self.mass_source_loads
                                       if e['lc'] == sl['lc']), 1.0)
                        m_imp += abs(sl['load']) * area * factor / GRAVITY

                m_shell = m_sw + m_imp
                cx, cy  = shell_centroid.get((area_label, story), (0.0, 0.0))
                cx_sum  += cx * m_shell
                cy_sum  += cy * m_shell
                total_mass += m_shell
                if area > 1e-9:
                    side = math.sqrt(area)
                    mmi += m_shell * side ** 2 / 6.0

            if total_mass > 1e-9:
                cx_cm = cx_sum / total_mass
                cy_cm = cy_sum / total_mass
            else:
                cx_cm = cy_cm = 0.0

            rows.append({
                'Story':                  story,
                'Diaphragm':              f'D{i}',
                'Mass X':                 round(total_mass, 4),
                'Mass Y':                 round(total_mass, 4),
                'Mass Moment of Inertia': round(mmi, 4),
                'X Mass Center':          round(cx_cm, 4),
                'Y Mass Center':          round(cy_cm, 4),
            })
        return pd.DataFrame(rows)

    def _build_restraints(self, df_joints: pd.DataFrame) -> pd.DataFrame:
        base_story = self.story_order[0] if self.story_order else ''
        rows = []
        for _, row in df_joints.iterrows():
            if row['Story'] == base_story:
                rows.append({
                    'Story': row['Story'],
                    'Unique Name': row['Element Label'],
                    'UX': 'Yes', 'UY': 'Yes', 'UZ': 'Yes',
                    'RX': 'Yes', 'RY': 'Yes', 'RZ': 'Yes',
                })
        return pd.DataFrame(rows)

    def _build_frame_sections(self) -> pd.DataFrame:
        rows = [
            {'Name': n, 'Material': s['Material'], 'Shape': s['Shape'],
             't3': s['t3'], 't2': s['t2'], 'Area': s['Area'],
             'I33': s['I33'], 'I22': s['I22'], 'J': s['J']}
            for n, s in self.frame_sec_defs.items()
        ]
        return pd.DataFrame(rows)

    def _build_frame_assignments(self, df_frames: pd.DataFrame) -> pd.DataFrame:
        rows = []
        for _, row in df_frames.iterrows():
            ln = row['Object Label']
            story = row['Story']
            sec = self.line_assigns.get((ln, story), {}).get('section', '')
            rows.append({
                'Story': story,
                'Label': ln,
                'Unique Name': row['Element Label'],
                'Analysis Section': sec,
                'Design Section': sec,
            })
        return pd.DataFrame(rows)

    def _build_shell_sections_slab(self) -> pd.DataFrame:
        rows = [
            {'Name': n, 'Material': s['Material'],
             'Slab Thickness': s['Thickness'],
             'One-Way Load Distribution?': s.get('OneWay', 'No')}
            for n, s in self.slab_props.items()
        ]
        return pd.DataFrame(rows)

    def _build_shell_sections_wall(self):
        if not self.wall_props:
            return None
        rows = [
            {'Name': n, 'Material': w['Material'], 'Thickness': w['Thickness']}
            for n, w in self.wall_props.items()
        ]
        return pd.DataFrame(rows)

    def _build_shell_assignments(self, df_shells: pd.DataFrame) -> pd.DataFrame:
        rows = []
        for _, row in df_shells.iterrows():
            ar    = row['Area Label']
            story = row['Story']
            sec   = self.area_assigns.get((ar, story), {}).get('section', '')
            rows.append({
                'Story': story,
                'Unique Name': row['Element Label'],
                'Section': sec,
            })
        return pd.DataFrame(rows)

    def _build_frame_loads_df(self) -> pd.DataFrame:
        rows = []
        for fl in self.frame_loads:
            ln    = fl['frame']
            story = fl['story']
            label = self._frame_elem_label.get((ln, story), 0)
            rows.append({
                'Story': story, 'Label': ln, 'Unique Name': label,
                'Load Pattern': fl['lc'], 'Direction': fl['direction'],
                'Distance A': 0.0, 'Distance B': 1.0,
                'Load A': fl['fval'], 'Load B': fl['fval'],
            })
        return pd.DataFrame(rows)

    def _build_shell_loads_df(self) -> pd.DataFrame:
        rows = []
        for sl in self.shell_loads:
            ar    = sl['area']
            story = sl['story']
            label = self._shell_elem_label.get((ar, story), 0)
            rows.append({
                'Story': story, 'Label': ar, 'Unique Name': label,
                'Load Pattern': sl['lc'], 'Direction': sl['direction'],
                'Load': sl['load'],
            })
        return pd.DataFrame(rows)

    def _build_column_rebar_data(self):
        rows = []
        for name, cs in self.concrete_secs.items():
            if cs['type'].lower() != 'column':
                continue
            rows.append({
                'Frame Property': name,
                '# Long. Bars 3-axis': cs['bars_3axis'],
                '# Long. Bars 2-axis': cs['bars_2axis'],
                'Corner Bar Area': round(cs['longbararea_m2'] * 1e6, 2),
                'Cover': cs['cover'],
            })
        return pd.DataFrame(rows) if rows else None

    def _build_pier_section_properties(self) -> pd.DataFrame:
        rows = [
            {'Story': story, 'Pier': pier, 'Length': 0.0, 'Thickness': 0.2}
            for pier in self.pier_names
            for story in self.story_order[1:]
        ]
        return pd.DataFrame(rows)

    def _build_shell_pier_assignments(self, df_shells: pd.DataFrame) -> pd.DataFrame:
        rows = [
            {'Story': row['Story'], 'Unique Name': row['Element Label'], 'Pier': ''}
            for _, row in df_shells.iterrows()
            if str(row.get('Area Type', '')).lower() == 'wall'
        ]
        return pd.DataFrame(rows)

    # ── supplemental merge ────────────────────────────────────────────────────

    def _merge_supplemental_sheets(self, wb: Workbook,
                                    supplemental_path: str) -> None:
        COPY_SHEETS = {
            'Concrete Beam Summary - ACI 318',
            'Concrete Column Summary - ACI 3',
            'Shear Wall Pier Summary - ACI 3',
            'Concrete Beam Rebar Data',
            'Concrete Column Rebar Data',
        }
        try:
            src_wb = load_workbook(supplemental_path, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"Cannot open supplemental XLSX: {e}")

        for sheet_name in src_wb.sheetnames:
            if sheet_name not in COPY_SHEETS:
                continue
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            src_ws = src_wb[sheet_name]
            dst_ws = wb.create_sheet(title=sheet_name)
            for row in src_ws.iter_rows(values_only=True):
                dst_ws.append(list(row))
        src_wb.close()
